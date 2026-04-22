"""
Coda Knowledge Engine V6.0 — Shadow Mirror + Ripple Update Engine
影子镜像: 二进制文件 → 语义化 Markdown
涟漪更新: 源文件变更 → 级联更新受影响页面

影子镜像 (Shadow Mirror):
  - MIME 路由: 根据文件类型选择提取策略
  - PDF → Markdown (结构保留)
  - Excel/CSV → Markdown 表格
  - 图片 → 描述 + alt-text
  - 音视频 → 元数据提取

涟漪更新 (Ripple Update):
  - 源文件变更时, 追踪所有引用该源的知识页面
  - 仅重编译受影响的子集 (非全量)
  - 承重边优先更新策略
"""

from __future__ import annotations

from ..base_plugin import WikiPlugin, WikiHook, WikiPluginContext

import hashlib
import json
import logging
import re
import time
from pathlib import Path
from typing import Any

from ..akp_types import (
    KnowledgeNode, KnowledgeRelation, RelationType, NodeStatus,
)
from .atlas import AtlasIndex
from .storage import CompilationManifest, WikiStorage

logger = logging.getLogger("Coda.wiki.shadow")


# ════════════════════════════════════════════
#  MIME 路由表
# ════════════════════════════════════════════

MIME_EXTRACTORS: dict[str, str] = {
    # 文本
    "text/markdown": "passthrough",
    "text/plain": "passthrough",
    "text/x-python": "code_extractor",
    "text/javascript": "code_extractor",
    "text/typescript": "code_extractor",
    "text/yaml": "passthrough",
    "text/csv": "csv_extractor",
    # 结构化数据
    "application/json": "json_extractor",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "excel_extractor",
    # 文档
    "application/pdf": "pdf_extractor",
    # 媒体
    "image/png": "image_extractor",
    "image/jpeg": "image_extractor",
    "image/gif": "image_extractor",
    "video/mp4": "media_metadata_extractor",
    "audio/mpeg": "media_metadata_extractor",
}


class ShadowMirror(WikiPlugin):
    """
    影子镜像: 将非 Markdown 文件转换为语义化的 .md 摘要。
    """
    name = "shadow"

    def __init__(
        self,
        storage: WikiStorage | None = None,
        manifest: CompilationManifest | None = None,
        llm_caller: Any = None,
    ):
        self._storage = storage
        self._manifest = manifest
        self._llm = llm_caller

    async def initialize(self, ctx: WikiPluginContext) -> None:
        """插件初始化入口。"""
        if not self._storage:
            self._storage = ctx.storage
        if not self._manifest:
            if hasattr(self._storage, "get_manifest"):
                self._manifest = self._storage.get_manifest()
        logger.info("🪞 Shadow Mirror plugin initialized")

    async def on_hook(self, hook: WikiHook, payload: Any = None) -> Any:
        """响应 Wiki 钩子。"""
        return None

    def mirror_file(self, filepath: Path) -> Path | None:
        """
        对单个文件执行影子镜像。
        返回生成的 .md 文件路径, 如果无需镜像则返回 None。
        """
        if filepath.suffix.lower() == ".md":
            return None  # Markdown 不需要镜像

        mime = self._guess_mime(filepath)
        extractor = MIME_EXTRACTORS.get(mime, "generic_extractor")

        logger.info(f"🪞 Shadow mirror: {filepath.name} ({mime}) → {extractor}")

        try:
            markdown_content = self._dispatch_extractor(extractor, filepath, mime)
        except Exception as e:
            logger.error(f"Shadow mirror failed for {filepath}: {e}")
            return None

        if not markdown_content or len(markdown_content.strip()) < 10:
            logger.warning(f"Shadow mirror produced empty output for {filepath.name}")
            return None

        # 写入 _meta/extracted/
        output_name = f"{filepath.stem}.md"
        output_dir = self._storage._root / "_meta" / "extracted"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / output_name

        # 生成 frontmatter
        content_hash = hashlib.sha256(filepath.read_bytes()).hexdigest()[:16]
        full_content = self._build_shadow_markdown(
            filepath, mime, markdown_content, content_hash
        )

        output_path.write_text(full_content, encoding="utf-8")

        # 注册到 Manifest
        self._manifest.mark_compiled(filepath, derived_pages=[filepath.stem])

        logger.info(f"✅ Shadow mirror created: {output_path.name}")
        return output_path

    def mirror_directory(self, source_dir: Path) -> list[Path]:
        """批量镜像目录下的所有非 Markdown 文件。"""
        mirrored: list[Path] = []
        if not source_dir.exists():
            return mirrored

        for filepath in source_dir.rglob("*"):
            if filepath.is_file() and filepath.suffix.lower() != ".md":
                if not filepath.name.startswith("."):
                    result = self.mirror_file(filepath)
                    if result:
                        mirrored.append(result)

        return mirrored

    def _dispatch_extractor(self, extractor: str, filepath: Path, mime: str) -> str:
        """根据提取器名称路由到具体实现。"""
        dispatch = {
            "passthrough": self._extract_passthrough,
            "code_extractor": self._extract_code,
            "csv_extractor": self._extract_csv,
            "json_extractor": self._extract_json,
            "excel_extractor": self._extract_excel,
            "pdf_extractor": self._extract_pdf,
            "image_extractor": self._extract_image,
            "media_metadata_extractor": self._extract_media_metadata,
            "generic_extractor": self._extract_generic,
        }
        handler = dispatch.get(extractor, self._extract_generic)
        return handler(filepath)

    # ── 各类提取器 ──

    def _extract_passthrough(self, filepath: Path) -> str:
        return filepath.read_text(encoding="utf-8", errors="replace")

    def _extract_code(self, filepath: Path) -> str:
        """提取代码文件的结构化摘要。"""
        content = filepath.read_text(encoding="utf-8", errors="replace")
        lang = filepath.suffix.lstrip(".")

        lines = content.split("\n")
        imports: list[str] = []
        classes: list[str] = []
        functions: list[str] = []
        docstrings: list[str] = []

        for line in lines:
            stripped = line.strip()
            if stripped.startswith(("import ", "from ")):
                imports.append(stripped)
            elif stripped.startswith("class "):
                classes.append(stripped.split("(")[0].split(":")[0].replace("class ", ""))
            elif stripped.startswith(("def ", "async def ")):
                func_name = stripped.replace("async ", "").split("(")[0].replace("def ", "")
                functions.append(func_name)
            elif stripped.startswith('"""') or stripped.startswith("'''"):
                docstrings.append(stripped.strip("\"'"))

        sections: list[str] = []
        sections.append(f"# Code Analysis: {filepath.name}\n")
        sections.append(f"**Language**: {lang}")
        sections.append(f"**Lines**: {len(lines)}\n")

        if classes:
            sections.append("## Classes\n" + "\n".join(f"- `{c}`" for c in classes))
        if functions:
            sections.append("## Functions\n" + "\n".join(f"- `{f}()`" for f in functions[:50]))
        if docstrings:
            sections.append("## Docstrings\n" + "\n".join(f"> {d}" for d in docstrings[:10]))

        return "\n\n".join(sections)

    def _extract_csv(self, filepath: Path) -> str:
        """提取 CSV 文件为 Markdown 表格。"""
        import csv
        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = list(reader)
        except Exception:
            return f"# CSV: {filepath.name}\n\n_Failed to parse CSV._"

        if not rows:
            return f"# CSV: {filepath.name}\n\n_Empty file._"

        header = rows[0]
        data_rows = rows[1:50]  # 限制行数

        result = [f"# CSV Data: {filepath.name}\n"]
        result.append(f"**Rows**: {len(rows) - 1}, **Columns**: {len(header)}\n")
        result.append("| " + " | ".join(header) + " |")
        result.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in data_rows:
            padded = row + [""] * (len(header) - len(row))
            result.append("| " + " | ".join(padded[:len(header)]) + " |")

        if len(rows) > 51:
            result.append(f"\n_...and {len(rows) - 51} more rows._")

        return "\n".join(result)

    def _extract_json(self, filepath: Path) -> str:
        """提取 JSON 文件的结构摘要。"""
        try:
            data = json.loads(filepath.read_text(encoding="utf-8"))
        except Exception as e:
            return f"# JSON: {filepath.name}\n\n_Parse error: {e}_"

        result = [f"# JSON Structure: {filepath.name}\n"]
        result.append(f"**Type**: {type(data).__name__}")

        if isinstance(data, dict):
            result.append(f"**Keys**: {len(data)}\n")
            result.append("## Top-level Keys\n")
            for key in list(data.keys())[:30]:
                val = data[key]
                val_type = type(val).__name__
                preview = str(val)[:100] if not isinstance(val, (dict, list)) else f"({val_type})"
                result.append(f"- `{key}`: {preview}")
        elif isinstance(data, list):
            result.append(f"**Items**: {len(data)}")
            if data:
                result.append(f"\n**First item type**: {type(data[0]).__name__}")

        return "\n".join(result)

    def _extract_excel(self, filepath: Path) -> str:
        """提取 Excel 文件 (需要 openpyxl)。"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except ImportError:
            return f"# Excel: {filepath.name}\n\n_openpyxl not installed._"
        except Exception as e:
            return f"# Excel: {filepath.name}\n\n_Parse error: {e}_"

        result = [f"# Excel: {filepath.name}\n"]
        result.append(f"**Sheets**: {', '.join(wb.sheetnames)}\n")

        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            result.append(f"## Sheet: {sheet_name}\n")
            rows_data = list(ws.iter_rows(max_row=20, values_only=True))
            if rows_data:
                header = [str(c) if c else "" for c in rows_data[0]]
                result.append("| " + " | ".join(header) + " |")
                result.append("| " + " | ".join(["---"] * len(header)) + " |")
                for row in rows_data[1:]:
                    cells = [str(c) if c else "" for c in row]
                    result.append("| " + " | ".join(cells[:len(header)]) + " |")

        wb.close()
        return "\n".join(result)

    def _extract_pdf(self, filepath: Path) -> str:
        """提取 PDF 文件 (需要 PyMuPDF 或 pdfplumber)。"""
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(filepath)
            result = [f"# PDF: {filepath.name}\n"]
            result.append(f"**Pages**: {len(doc)}\n")
            for page_num in range(min(len(doc), 20)):
                page = doc[page_num]
                text = page.get_text().strip()
                if text:
                    result.append(f"## Page {page_num + 1}\n")
                    result.append(text[:2000])
            doc.close()
            return "\n".join(result)
        except ImportError:
            pass

        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                result = [f"# PDF: {filepath.name}\n"]
                result.append(f"**Pages**: {len(pdf.pages)}\n")
                for page in pdf.pages[:20]:
                    text = page.extract_text() or ""
                    if text.strip():
                        result.append(f"## Page {page.page_number}\n")
                        result.append(text[:2000])
                return "\n".join(result)
        except ImportError:
            return f"# PDF: {filepath.name}\n\n_No PDF library available (install PyMuPDF or pdfplumber)._"

    def _extract_image(self, filepath: Path) -> str:
        """提取图片元数据。"""
        result = [f"# Image: {filepath.name}\n"]
        stat = filepath.stat()
        result.append(f"**Size**: {stat.st_size / 1024:.1f} KB")
        result.append(f"**Format**: {filepath.suffix.upper()}")

        try:
            from PIL import Image
            img = Image.open(filepath)
            result.append(f"**Dimensions**: {img.width} × {img.height}")
            result.append(f"**Mode**: {img.mode}")
            if img.info:
                result.append("\n## EXIF/Metadata")
                for k, v in list(img.info.items())[:10]:
                    result.append(f"- `{k}`: {str(v)[:200]}")
            img.close()
        except ImportError:
            result.append("\n_Pillow not installed for full image analysis._")

        return "\n".join(result)

    def _extract_media_metadata(self, filepath: Path) -> str:
        """提取音视频元数据。"""
        result = [f"# Media: {filepath.name}\n"]
        stat = filepath.stat()
        result.append(f"**Size**: {stat.st_size / (1024*1024):.2f} MB")
        result.append(f"**Format**: {filepath.suffix.upper()}")
        return "\n".join(result)

    def _extract_generic(self, filepath: Path) -> str:
        """通用提取器 (仅元数据)。"""
        stat = filepath.stat()
        return (
            f"# File: {filepath.name}\n\n"
            f"**Size**: {stat.st_size / 1024:.1f} KB\n"
            f"**Type**: {filepath.suffix}\n"
        )

    def _build_shadow_markdown(self, source: Path, mime: str,
                               content: str, content_hash: str) -> str:
        """构建带 frontmatter 的影子 Markdown。"""
        frontmatter = (
            f"---\n"
            f"id: shadow-{source.stem}\n"
            f"title: \"Shadow: {source.name}\"\n"
            f"type: source\n"
            f"status: validated\n"
            f"epistemic_tag: confirmed\n"
            f"source_origin_hash: {content_hash}\n"
            f"pii_shield: raw\n"
            f"---\n\n"
        )
        return frontmatter + content

    def _guess_mime(self, path: Path) -> str:
        """基于扩展名推断 MIME 类型。"""
        mime_map = {
            ".md": "text/markdown", ".txt": "text/plain",
            ".py": "text/x-python", ".js": "text/javascript",
            ".ts": "text/typescript", ".json": "application/json",
            ".yaml": "text/yaml", ".yml": "text/yaml",
            ".csv": "text/csv",
            ".pdf": "application/pdf",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".png": "image/png", ".jpg": "image/jpeg", ".gif": "image/gif",
            ".mp4": "video/mp4", ".mp3": "audio/mpeg",
        }
        return mime_map.get(path.suffix.lower(), "application/octet-stream")


# ════════════════════════════════════════════
#  Ripple Update Engine
# ════════════════════════════════════════════

class RippleEngine(WikiPlugin):
    """
    涟漪更新引擎 — 当源文件变更时, 追踪并重编译受影响的知识页面。
    """
    name = "ripple"

    def __init__(
        self,
        atlas: AtlasIndex | None = None,
        manifest: CompilationManifest | None = None,
        storage: WikiStorage | None = None,
    ):
        self._atlas = atlas
        self._manifest = manifest
        self._storage = storage

    async def initialize(self, ctx: WikiPluginContext) -> None:
        """插件初始化入口。"""
        if not self._atlas:
            self._atlas = ctx.atlas
        if not self._storage:
            self._storage = ctx.storage
        if not self._manifest:
             if hasattr(self._storage, "get_manifest"):
                self._manifest = self._storage.get_manifest()
        logger.info("🌊 Ripple Engine plugin initialized")

    async def on_hook(self, hook: WikiHook, payload: Any = None) -> Any:
        """响应 Wiki 钩子。"""
        return None

    def compute_affected_set(self, changed_source: str | Path) -> list[str]:
        """
        计算受影响的页面集合 (Affected Sources Cascade)。

        Returns: 需要重编译的 node ID 列表 (按承重优先排序)。
        """
        # Step 1: 直接受影响的页面
        direct_pages = self._manifest.get_affected_pages(str(changed_source))
        if not direct_pages:
            return []

        # Step 2: 级联追踪
        all_affected: set[str] = set(direct_pages)
        load_bearing_set: set[str] = set()
        frontier = list(direct_pages)
        depth = 0

        while frontier and depth < self.MAX_CASCADE_DEPTH:
            next_frontier: list[str] = []
            for page_id in frontier:
                # 查找所有引用此页面的节点
                neighbors = self._atlas.get_neighbors(page_id, max_hops=1)
                for neighbor in neighbors:
                    nid = neighbor["node_id"]
                    if nid not in all_affected:
                        all_affected.add(nid)
                        next_frontier.append(nid)
                        if neighbor.get("load_bearing"):
                            load_bearing_set.add(nid)

            frontier = next_frontier
            depth += 1

        # Step 3: 排序 (承重边优先)
        result = sorted(all_affected, key=lambda x: x in load_bearing_set, reverse=True)
        logger.info(
            f"🌊 Ripple: {len(direct_pages)} direct → {len(result)} total "
            f"(depth={depth}, load_bearing={len(load_bearing_set)})"
        )
        return result

    def execute_ripple(self, changed_source: str | Path,
                       recompile_fn: Any = None) -> dict[str, Any]:
        """
        执行涟漪更新。

        Args:
            changed_source: 变更的源文件路径
            recompile_fn: 重编译函数 (接收 node_id, 返回 bool)

        Returns: 更新报告。
        """
        affected = self.compute_affected_set(changed_source)

        stats = {
            "source": str(changed_source),
            "affected_count": len(affected),
            "recompiled": 0,
            "skipped": 0,
            "errors": 0,
        }

        for node_id in affected:
            if recompile_fn:
                try:
                    success = recompile_fn(node_id)
                    if success:
                        stats["recompiled"] += 1
                    else:
                        stats["skipped"] += 1
                except Exception as e:
                    logger.error(f"Ripple recompile failed for {node_id}: {e}")
                    stats["errors"] += 1
            else:
                stats["skipped"] += 1  # 无重编译函数, 仅标记

        self._storage.append_audit_log(
            action="ripple_update",
            target=str(changed_source),
            detail=f"Affected: {len(affected)}, Recompiled: {stats['recompiled']}",
        )

        return stats

    def find_shared_concept_cascade(self, concept_id: str) -> list[str]:
        """
        概念级联: 查找所有引用同一概念的源文件。
        即使源文件 B 未变, 如果概念 X 的定义改了, B 也需要重编。
        """
        sources = self._manifest.find_shared_concept_sources(concept_id)
        return sources
